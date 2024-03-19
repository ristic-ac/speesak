import os
from openpyxl import load_workbook
import pandas as pd

def xlsx_to_array(workbook_path):
    # Load the workbook
    workbook = load_workbook(workbook_path)

    # Keep track of the number of rows
    row_no = 1

    # Placeholder for the data
    data = []

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        # Get the sheet
        sheet = workbook[sheet_name]

        # If sheet name is "Worksheet", skip it
        if sheet_name == "Worksheet":
            continue

        # Sheet name: Grupa 13 -> 13
        group = sheet_name.split(" ")[1]

        # Extract data from the sheet
        for row in sheet.iter_rows(min_row=7, values_only=True):
            # Replace first field in row with row_no
            row = (row_no,) + row[1:]

            # Add group number to the row
            row = (group,) + row

            # Append the row to the data
            data.append(row)

            # Increment the row number
            row_no += 1

    df = pd.DataFrame(data, columns=["Grupa", "Redni_broj", "Broj indeksa", "Prezime", "Ime"])
    return df


def xlsx_k_to_array(workbook_path):
    df = pd.read_excel(workbook_path, sheet_name='dhtmlxGrid')
    # Keep columns: "Broj indeksa", "Prezime", "Ime", "Način slušanja"
    df = df[["Broj indeksa", "Prezime", "Ime", "Način slušanja"]]
    print(df.columns)
    return df

        

# Example usage:
xlsx_file = "./xlsx/RA.xlsx"
xlsx_k_file = "./xlsx/RAK.xlsx"

df = xlsx_to_array(xlsx_file)
print(df)
df_k = xlsx_k_to_array(xlsx_k_file)
print(df_k)

# Find intersection based on "Broj indeksa"
df_merged = pd.merge(df, df_k, on="Broj indeksa", how="inner")
print(df_merged)

# Find 'Broj indeksa' that are in df but not in df_merged
df_not_merged = df[~df['Broj indeksa'].isin(df_merged['Broj indeksa'])]
print(df_not_merged)
