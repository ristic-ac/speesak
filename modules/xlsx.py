import os
from openpyxl import load_workbook
import pandas as pd

def convert_xlsx_groups_to_dataframe(xlsx_file):
    """
    Pretvara podatke iz više grupa (radnih listova) u XLSX fajlu u pandas DataFrame.

    Svaki radni list (osim "Worksheet") predstavlja jednu grupu, pri čemu se identifikator grupe
    izvlači iz imena radnog lista (druga reč). Podaci se čitaju počevši od 7. reda svakog radnog lista.
    Svaki red u rezultujućem DataFrame-u sadrži grupu, redni broj i izabrane kolone iz XLSX fajla.

    Argumenti:
        xlsx_file (str): Naziv XLSX fajla koji se nalazi u direktorijumu "xlsx".

    Povratna vrednost:
        pandas.DataFrame: DataFrame sa kolonama ["Grupa", "Redni_broj", "Broj indeksa", "Prezime", "Ime"],
        gde svaki red odgovara jednom studentu iz grupa.
    """
    xlsx_file = os.path.join("xlsx", xlsx_file)
    workbook = load_workbook(xlsx_file)
    row_no = 1
    data = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet_name == "Worksheet":
            continue
        group = sheet_name.split(" ")[1]

        for row in sheet.iter_rows(min_row=7, values_only=True):
            row = (row_no,) + row[1:]
            row = (group,) + row
            data.append(row)
            row_no += 1
    return pd.DataFrame(data, columns=["Grupa", "Redni_broj", "Broj indeksa", "Prezime", "Ime"])

def convert_xlsx_complete_to_dataframe(workbook_path):
    """
    Pretvara navedeni XLSX fajl u pandas DataFrame, izdvajajući određene kolone.

    Argumenti:
        workbook_path (str): Naziv XLSX fajla koji se nalazi u poddirektorijumu 'xlsx'.

    Povratna vrednost:
        pandas.DataFrame: DataFrame sa kolonama 'Broj indeksa', 'Prezime', 'Ime' i 'Način slušanja' iz radnog lista 'dhtmlxGrid'.

    Izuzeci:
        FileNotFoundError: Ako navedeni XLSX fajl ne postoji.
        ValueError: Ako tražene kolone nisu pronađene u radnom listu.
        ImportError: Ako pandas nije instaliran.
    """
    workbook_path = os.path.join("xlsx", workbook_path)
    df = pd.read_excel(workbook_path, sheet_name='dhtmlxGrid')
    df = df[["Broj indeksa", "Prezime", "Ime", "Način slušanja"]]
    return df